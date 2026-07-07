"""Tests d'intégration : scénario de compromission et scénario bénin sur fixtures,
plus tests sur vrais logs (marqués slow)."""
import os
import time
import pytest
import shutil
import tempfile
import yaml
from pathlib import Path
from tests.conftest import (detect_on_fixture, require_real_logs, FIXTURES, CONFIG_PATH,
                            REAL_LOGS_T1, REAL_LOGS_T2)


# --- Tests sur fixtures synthétiques ---

def test_compromission_scenario_critiques(cfg):
    """Le scénario de compromission doit remonter au moins 3 événements critiques."""
    ev = detect_on_fixture("compromission_scenario.log", cfg)
    crit = ev[ev["severite"] == "critique"]
    assert len(crit) >= 3, f"Attendu ≥3 critiques, trouvé {len(crit)}: {crit['regle'].tolist()}"


def test_compromission_scenario_eleves(cfg):
    """Le scénario de compromission doit aussi remonter des événements élevés."""
    ev = detect_on_fixture("compromission_scenario.log", cfg)
    eleve = ev[ev["severite"] == "eleve"]
    assert len(eleve) >= 1


def test_benign_zero_faux_positifs_voyou(cfg):
    """Les logs bénins ne doivent PAS générer de faux positif sur les comptes voyous.
    Un nom suspect (admin-1) dans un login ÉCHOUÉ ne doit pas être signalé comme voyou."""
    ev = detect_on_fixture("benign_mix.log", cfg)
    rogue = ev[ev["regle"].str.contains("voyou", na=False)]
    assert rogue.empty, f"Faux positif voyou : {rogue['detail'].tolist()}"


def test_benign_no_critique(cfg):
    """Les logs bénins ne doivent pas générer de critique."""
    ev = detect_on_fixture("benign_mix.log", cfg)
    crit = ev[ev["severite"] == "critique"]
    assert crit.empty, f"Faux positif critique : {crit['regle'].tolist()}"


def test_run_end_to_end_with_chains():
    """run() complet sur la fixture de compromission : table chains + feuille Excel."""
    from fortilog.main import run
    from openpyxl import load_workbook

    input_dir = Path(tempfile.mkdtemp())
    output_dir = Path(tempfile.mkdtemp())
    try:
        shutil.copy(FIXTURES / "compromission_scenario.log", input_dir / "scenario.log")
        tables, meta = run(str(input_dir), str(CONFIG_PATH), str(output_dir))

        assert "chains" in tables
        assert not tables["chains"].empty
        assert (tables["chains"]["severite"] == "critique").any()

        xlsx = output_dir / "rapport_fortigate.xlsx"
        assert xlsx.exists()
        wb = load_workbook(xlsx)
        assert "Chaines suspectes" in wb.sheetnames
    finally:
        shutil.rmtree(input_dir, ignore_errors=True)
        shutil.rmtree(output_dir, ignore_errors=True)


def test_run_end_to_end_with_geo():
    """run() complet : portée srcip enrichie + feuille Excel 'Sources externes'.
    Sans base géo configurée -> la portée reste calculée, la feuille existe."""
    from fortilog.main import run
    from openpyxl import load_workbook

    input_dir = Path(tempfile.mkdtemp())
    output_dir = Path(tempfile.mkdtemp())
    try:
        shutil.copy(FIXTURES / "login_admin_ext.log", input_dir / "ext.log")
        tables, meta = run(str(input_dir), str(CONFIG_PATH), str(output_dir))

        assert "sources_externes" in tables
        assert "geo_available" in meta
        ev = tables["events"]
        if not ev.empty:
            assert "srcip_portee" in ev.columns

        wb = load_workbook(output_dir / "rapport_fortigate.xlsx")
        assert "Sources externes" in wb.sheetnames
        assert "IP malveillantes" in wb.sheetnames
        assert "Rapport" in wb.sheetnames
        assert "reputation" in tables
        assert "reputation_available" in meta
        assert "analysis" in meta and "# RAPPORT D'ANALYSE" in meta["analysis"]
    finally:
        shutil.rmtree(input_dir, ignore_errors=True)
        shutil.rmtree(output_dir, ignore_errors=True)


def test_run_config_only():
    """Import de .conf SANS logs : run() produit l'audit config + feuille Excel."""
    from fortilog.main import run
    from openpyxl import load_workbook

    input_dir = Path(tempfile.mkdtemp())
    output_dir = Path(tempfile.mkdtemp())
    try:
        shutil.copy(FIXTURES / "confaudit_compromis.conf", input_dir / "fw.conf")
        tables, meta = run(str(input_dir), str(CONFIG_PATH), str(output_dir))

        assert "config_audit" in tables
        assert not tables["config_audit"].empty
        assert (tables["config_audit"]["severite"] == "critique").any()
        assert meta["n_configs"] == 1
        assert tables["config_diff"].empty   # pas de config de référence -> pas de comparaison

        wb = load_workbook(output_dir / "rapport_fortigate.xlsx")
        assert "Audit config" in wb.sheetnames
        assert "Rapport" in wb.sheetnames
    finally:
        shutil.rmtree(input_dir, ignore_errors=True)
        shutil.rmtree(output_dir, ignore_errors=True)


def test_run_with_config_diff():
    """run() avec une config de RÉFÉRENCE : la comparaison est intégrée au rapport global."""
    from fortilog.main import run
    from openpyxl import load_workbook

    input_dir = Path(tempfile.mkdtemp())
    output_dir = Path(tempfile.mkdtemp())
    try:
        shutil.copy(FIXTURES / "login_admin_ext.log", input_dir / "ev.log")
        shutil.copy(FIXTURES / "confdiff_current.conf", input_dir / "current.conf")
        ref = FIXTURES / "confdiff_ok.conf"
        tables, meta = run(str(input_dir), str(CONFIG_PATH), str(output_dir), ref_conf=str(ref))

        cd = tables["config_diff"]
        assert not cd.empty
        assert (cd["objet"] == "backdoor").any()        # compte ajouté détecté
        assert meta["n_config_changes"] > 0
        assert meta["config_ref"] == "confdiff_ok.conf"
        assert "Changements de configuration" in meta["analysis"]

        wb = load_workbook(output_dir / "rapport_fortigate.xlsx")
        assert "Comparaison config" in wb.sheetnames
    finally:
        shutil.rmtree(input_dir, ignore_errors=True)
        shutil.rmtree(output_dir, ignore_errors=True)


# --- Tests sur vrais logs (marqués slow) ---

@pytest.mark.slow
def test_real_logs_t1_parseable(cfg):
    """Vérifie que les vrais logs T1 se parsent sans erreur."""
    from fortilog.main import load_file
    from fortilog import ingest
    require_real_logs(REAL_LOGS_T1)
    files = ingest.list_log_files(REAL_LOGS_T1)
    assert len(files) > 0
    for f in files:
        t, s, reconnu = ingest.detect_type(f)
        assert t != "", f"Type vide pour {f.name}"


@pytest.mark.slow
def test_real_logs_t2_parseable(cfg):
    """Vérifie que les vrais logs T2 se parsent sans erreur."""
    from fortilog.main import load_file
    from fortilog import ingest
    require_real_logs(REAL_LOGS_T2)
    files = ingest.list_log_files(REAL_LOGS_T2)
    assert len(files) > 0
    for f in files:
        t, s, reconnu = ingest.detect_type(f)
        assert t != "", f"Type vide pour {f.name}"


@pytest.mark.slow
def test_real_event_system_t1_detection(cfg):
    """Vérifie la détection sur le vrai fichier event/system T1."""
    from fortilog.main import load_file
    from fortilog import normalize, detect

    require_real_logs(REAL_LOGS_T1)
    f = REAL_LOGS_T1 / "memory-event-system-2026_06_23.log"
    df = load_file(f)
    assert len(df) > 200000
    df["timestamp"] = normalize.build_timestamp(df)
    df["boitier"] = normalize.assign_boitier(df, cfg.get("boitiers", {}), cfg.get("fichiers_boitier"))
    df = normalize.deduplicate(df)
    ev = detect.run_detection(df, cfg)
    assert len(ev) > 0
    assert "info" in ev["severite"].values


@pytest.mark.slow
def test_real_app_ctrl_t1_whitelist(cfg):
    """Sur les vrais logs app-ctrl T1, la whitelist élimine tous les faux positifs :
    proxy-safebrowsing.googleapis.com (Safe Browsing) ne doit pas déclencher R10b/c."""
    from fortilog.main import load_file
    from fortilog import normalize, detect

    require_real_logs(REAL_LOGS_T1)
    f = REAL_LOGS_T1 / "forticloud-app-ctrl-2026_06_23.log"
    df = load_file(f)
    df["timestamp"] = normalize.build_timestamp(df)
    df["boitier"] = normalize.assign_boitier(df, cfg.get("boitiers", {}), cfg.get("fichiers_boitier"))
    df = normalize.deduplicate(df)
    ev = detect.run_detection(df, cfg)
    # Avec la whitelist, aucun faux positif proxy-safebrowsing attendu
    r10bc = ev[ev["regle"].str.contains("risque critique|proxy", case=False, na=False)]
    false_pos = r10bc[r10bc["detail"].str.contains("proxy-safebrowsing", na=False)]
    assert false_pos.empty, f"Faux positif proxy-safebrowsing non filtré : {false_pos['detail'].tolist()}"


@pytest.mark.slow
def test_real_no_bruteforce_success_t2(cfg):
    """R11 sur le vrai event-system T2 (le plus attaqué) : aucun brute-force réussi —
    aucune source externe n'a percé, et l'admin interne légitime n'est pas un faux positif."""
    from fortilog.main import load_file
    from fortilog import normalize, detect

    require_real_logs(REAL_LOGS_T2)
    f = REAL_LOGS_T2 / "memory-event-system-2026_06_23.log"
    df = load_file(f)
    df["timestamp"] = normalize.build_timestamp(df)
    df["boitier"] = normalize.assign_boitier(df, cfg.get("boitiers", {}), cfg.get("fichiers_boitier"))
    df = normalize.deduplicate(df)
    ev = detect.run_detection(df, cfg)
    r11 = ev[ev["regle"].str.contains("potentiellement réussi|après rafale", na=False)]
    assert r11.empty, f"Brute-force réussi remonté sur vrai T2 : {r11['detail'].tolist()}"


@pytest.mark.slow
def test_real_geo_enrichment_t2(cfg):
    """Sur les vraies sources de brute-force T2, l'enrichissement géo/ASN (mini-base
    couvrant 85.11.x->GB) classe et contextualise les attaquants ; les plages hors
    base restent vides (jamais d'invention)."""
    from fortilog.main import load_file
    from fortilog import normalize, geo

    cfg2 = dict(cfg)
    cfg2["geo_db_path"] = str(FIXTURES / "geo_country_mini.csv")
    cfg2["asn_db_path"] = str(FIXTURES / "geo_asn_mini.tsv")

    require_real_logs(REAL_LOGS_T2)
    f = REAL_LOGS_T2 / "memory-event-system-2026_06_23.log"
    df = load_file(f)
    df["timestamp"] = normalize.build_timestamp(df)
    top = geo.top_external_sources(df, cfg2, n=10)
    assert not top.empty
    # le top attaquant est externe et enrichi par la mini-base
    row = top[top["srcip"] == "85.11.187.120"]
    assert not row.empty
    assert row.iloc[0]["srcip_portee"] == "externe"
    assert row.iloc[0]["srcip_pays"] == "GB"
    assert int(row.iloc[0]["logins_echoues"]) > 1000


@pytest.mark.slow
def test_real_config_audit_t1(cfg):
    """Audit du vrai .conf T1 : constats réels indépendants du référentiel —
    admins sans trusted-host + accès admin exposé sur WAN."""
    from fortilog import confaudit
    require_real_logs(REAL_LOGS_T1)
    confs = list(REAL_LOGS_T1.glob("FW-*.conf"))
    if not confs:
        pytest.skip("Pas de fichier .conf réel dans T1")
    df = confaudit.audit_files(confs, cfg)
    assert not df.empty
    assert df["regle"].str.contains("trusted-host").any()
    assert df["regle"].str.contains("WAN").any()


@pytest.mark.slow
def test_real_event_system_t1_no_false_chain(cfg):
    """Sur le vrai fichier event/system T1 (activité admin réelle), aucune fausse
    chaîne d'intrusion ne doit être remontée."""
    from fortilog.main import load_file
    from fortilog import normalize, detect, correlate

    require_real_logs(REAL_LOGS_T1)
    f = REAL_LOGS_T1 / "memory-event-system-2026_06_23.log"
    df = load_file(f)
    df["timestamp"] = normalize.build_timestamp(df)
    df["boitier"] = normalize.assign_boitier(df, cfg.get("boitiers", {}), cfg.get("fichiers_boitier"))
    df = normalize.deduplicate(df)
    ev = detect.run_detection(df, cfg)
    chains = correlate.correlate_chains(ev, cfg)
    assert chains.empty, f"Fausse(s) chaîne(s) sur vrai fichier T1 : {chains['detail'].tolist() if not chains.empty else ''}"


@pytest.mark.slow
def test_r13_campagne_massive_bornee(cfg):
    """La campagne name_invalid massive connue (T2, IP dense type 85.11.187.120) remonte
    en un nombre borné d'événements R13 (< 1000, un par IP×fenêtre — jamais un par ligne).
    NB : la campagne T1 est un spray distribué (~12 échecs/h max par IP), sous le seuil
    par IP — R13 y reste silencieux par conception (le volume reste visible via
    « Sources externes »)."""
    from fortilog.main import load_file
    from fortilog import normalize, detect
    require_real_logs(REAL_LOGS_T2)
    f = REAL_LOGS_T2 / "memory-event-system-2026_06_29.log"
    if not f.exists():
        pytest.skip(f"Fichier absent : {f}")
    df = load_file(f)
    df["timestamp"] = normalize.build_timestamp(df)
    df["boitier"] = normalize.assign_boitier(df, cfg.get("boitiers", {}), cfg.get("fichiers_boitier"))
    ev = detect.run_detection(df, cfg)
    r13 = ev[ev["regle"].str.contains("comptes inexistants", na=False)]
    assert 1 <= len(r13) < 1000, f"{len(r13)} événements R13"


def test_run_acteurs_et_frise_end_to_end():
    """run() complet : table acteurs remplie, feuille Excel « Acteurs a risque »,
    sections 3bis (acteurs) et FRISE CHRONOLOGIQUE dans le rapport de synthèse."""
    from fortilog.main import run
    from openpyxl import load_workbook

    input_dir = Path(tempfile.mkdtemp())
    output_dir = Path(tempfile.mkdtemp())
    try:
        shutil.copy(FIXTURES / "compromission_scenario.log", input_dir / "scenario.log")
        tables, meta = run(str(input_dir), str(CONFIG_PATH), str(output_dir))

        act = tables["acteurs"]
        assert not act.empty
        assert set(act["acteur_type"]) <= {"ip", "compte"}
        score_col = [c for c in act.columns if c.startswith("score_priorisation")][0]
        assert act[score_col].is_monotonic_decreasing

        assert "Acteurs à investiguer en priorité" in meta["analysis"]
        assert "FRISE CHRONOLOGIQUE" in meta["analysis"]

        wb = load_workbook(output_dir / "rapport_fortigate.xlsx")
        assert "Acteurs a risque" in wb.sheetnames
    finally:
        shutil.rmtree(input_dir, ignore_errors=True)
        shutil.rmtree(output_dir, ignore_errors=True)


def test_run_base_perimee_avertit_end_to_end():
    """run() complet avec une base géo vieillie : avertissement en tête de synthèse
    et ligne dédiée dans la feuille « Referentiel »."""
    from fortilog.main import run
    from openpyxl import load_workbook

    input_dir = Path(tempfile.mkdtemp())
    output_dir = Path(tempfile.mkdtemp())
    try:
        shutil.copy(FIXTURES / "compromission_scenario.log", input_dir / "scenario.log")

        geo_db = output_dir / "geo_vieille.csv"
        geo_db.write_text("start_ip,end_ip,country_code\n")
        ancien = time.time() - 100 * 86400
        os.utime(geo_db, (ancien, ancien))

        cfg = yaml.safe_load(CONFIG_PATH.read_text())
        cfg["geo_db_path"] = str(geo_db)
        cfg["bases"] = {"age_max_jours": 90}
        config_path = output_dir / "config_test.yaml"
        config_path.write_text(yaml.safe_dump(cfg))

        tables, meta = run(str(input_dir), str(config_path), str(output_dir))

        assert any(b["nom"] == "Géo (pays)" and b["perime"] for b in meta["bases"])
        assert "a 100 jours" in meta["analysis"]

        wb = load_workbook(output_dir / "rapport_fortigate.xlsx")
        ref_ws = wb["Referentiel"]
        cles = [row[0] for row in ref_ws.iter_rows(min_row=2, values_only=True)]
        assert any(c == "base:Géo (pays)" for c in cles)
    finally:
        shutil.rmtree(input_dir, ignore_errors=True)
        shutil.rmtree(output_dir, ignore_errors=True)
